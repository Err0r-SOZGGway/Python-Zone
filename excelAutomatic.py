import openpyxl
wb_matome=openpyxl.load_workbook('出社在宅集計表_まとめ.xlsx')
month_list=[]
for sheet_name in wb_matome.sheetnames:
    month_list.append(sheet_name)

busyo_list=['人事部','企画部','営業部']

for busyo in busyo_list:
    
    wb_kakubu=openpyxl.load_workbook('出社在宅集計表_{}.xlsx'.format(busyo))
        
    for month in month_list:
        ws_matome=wb_matome[month]
        ws_kakubu=wb_kakubu[month]
        
        for j in range(1,ws_matome.max_row+1):
            if ws_matome.cell(row=j,column=1).value==busyo:
             syussya_row=j
             zaitaku_row=j+1

        for i in range(ws_kakubu.max_column-1):
            ws_matome.cell(row=syussya_row,column=i+3).value=ws_kakubu.cell(row=2,column=i+2).value
            ws_matome.cell(row=zaitaku_row,column=i+3).value=ws_kakubu.cell(row=3,column=i+2).value


wb_matome.save('出社在宅集計表_まとめ5.xlsx')